import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import random
import pyperclip
import threading
import logging

names = []
city = []
name_button = None
city_button = None
attached_document = ''
sent_messages_count = 0
total_messages = 0

# Set up logging
logging.basicConfig(filename='message_sending.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

def preprocess_numbers(numbers_text):
    processed_numbers = []
    seen_numbers = set()
    for line in numbers_text.split('\n'):
        line = line.strip()
        if line:
            numbers = line.split()
            for num in numbers:
                if num not in seen_numbers:
                    processed_numbers.append(num)
                    seen_numbers.add(num)
    return processed_numbers

def read_excel(filename):
    global names, city
    try:
        wb = load_workbook(filename)
        sheet = wb.active
        numbers = []
        names = []
        city = []
        found_phone_column = False
        found_name_column = False
        found_city_column = False

        for row in sheet.iter_rows(values_only=True):
            if not found_phone_column:
                if "Number" in row:
                    phone_index = row.index("Number")
                    found_phone_column = True
            if not found_name_column:
                if "Name" in row:
                    name_index = row.index("Name")
                    found_name_column = True
            if not found_city_column:
                if 'City' in row:
                    city_index = row.index("City")
                    found_city_column = True

        for row in sheet.iter_rows(values_only=True):
            if row[phone_index] is not None:
                numbers.append(row[phone_index])
            if found_name_column:
                if row[name_index] is not None:
                    names.append(row[name_index])
                    if len(names) >= 2:
                        name_button.config(state=tk.NORMAL)
                else:
                    names.append('')
            if found_city_column:
                if row[city_index] is not None:
                    city.append(row[city_index])
                    if len(city) >= 2:
                        city_button.config(state=tk.NORMAL)
                else:
                    city.append('')
        if "Number" in numbers:
            numbers.remove('Number')
        if "Name" in names:
            names.remove("Name")
        if 'City' in city:
            city.remove('City')

        numbers_entry.delete("1.0", "end")
        numbers_entry.insert("1.0", "\n".join(map(str, numbers)))
        return names, city
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while reading Excel file: {e}")
        logging.error(f"Error reading Excel: {e}")

def upload_excel():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filename:
        global names, city
        names, city = read_excel(filename)
        delete_button.config(state=tk.NORMAL)

def insert_query_name():
    cursor_pos = message_entry.index(tk.INSERT)
    message_entry.insert(cursor_pos, '{{query_name}}')

def insert_query_city():
    cursor_pos = message_entry.index(tk.INSERT)
    message_entry.insert(cursor_pos, '{{query_city}}')

def attach_file():
    global attached_document
    attached_document = filedialog.askopenfilename(filetypes=[("Image Files", "*.png;*.jpg;*.jpeg;*.gif"), ("Video Files", "*.mp4;*.avi;*.mov")])
    attached_file_name = attached_document.split('/')[-1]
    attachment_label.config(text=f"{attached_file_name}")
    delete_attachment.config(state=tk.NORMAL)

def send_messages_async():
    threading.Thread(target=send_messages).start()

def send_messages():
    global names, city, sent_messages_count, total_messages
    numbers_text = numbers_entry.get("1.0", "end")
    original_message = message_entry.get("1.0", "end").strip()

    if not numbers_text or not original_message:
        messagebox.showerror("Error", "Please enter numbers and message")
        return
    numbers = preprocess_numbers(numbers_text)

    driver = None
    try:
        driver = webdriver.Chrome()
        driver.maximize_window()
        driver.get('https://web.whatsapp.com')
        WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, "//div[@title='New chat']")))

        batch_size = 30
        pause_duration = 60

        total_messages = len(numbers)
        sent_messages_count = 0

        for i in range(0, len(numbers), batch_size):
            batch_numbers = numbers[i:i + batch_size]
            if len(names) == 0 or len(city) == 0:
                for num in batch_numbers:
                    name = ''
                    city_name = ''
                    try:
                        send_message(driver, num, name, city_name, original_message)
                        sent_messages_count += 1
                        update_sent_messages_count(sent_messages_count, total_messages)
                        logging.info(f"Message sent to {num}")
                    except Exception as e:
                        logging.error(f"Error sending message to {num}: {e}")
                        continue
            else:
                for num, name, city_name in zip(batch_numbers, names, city):
                    try:
                        send_message(driver, num, name, city_name, original_message)
                        sent_messages_count += 1
                        update_sent_messages_count(sent_messages_count, total_messages)
                        logging.info(f"Message sent to {num} ({name}, {city_name})")
                    except Exception as e:
                        logging.error(f"Error sending message to {num}: {e}")
                        continue

            if i + batch_size < len(numbers):
                print(f"Sent messages to {i + batch_size} contacts. Taking a break for 1 minute...")
                time.sleep(pause_duration)
    except Exception as e:
        logging.error(f"An error occurred: {e}")
        print("An error occurred:", e)
    finally:
        if driver:
            time.sleep(2)
            driver.quit()
            messagebox.showinfo("Success", "Messages sent successfully")

def update_sent_messages_count(sent, total):
    sent_label.config(text=f"Messages sent: {sent}/{total}")
    progress_var.set((sent / total) * 100)  # Update progress bar

def send_message(driver, num, name, city_name, original_message):
    global attached_document
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@title='New chat']")))
    new_chat_button = driver.find_element(By.XPATH, "//div[@title='New chat']")
    new_chat_button.click()

    search_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@contenteditable='true' and @role='textbox' and @title='Search input textbox']")))
    search_input.send_keys(num)
    time.sleep(0.5)

    loading = driver.find_elements(By.XPATH, "//span[@class='_11JPr' and contains(text(), 'Looking outside your contacts...')]")
    if loading:
        try:
            WebDriverWait(driver, 100).until(EC.invisibility_of_element_located((By.XPATH, "//span[@class='_11JPr' and contains(text(), 'Looking outside your contacts...')]")))
        except TimeoutException:
            logging.warning(f"Loading element did not disappear for {num}")
            print("Loading element did not disappear within the timeout period")

    in_contact = driver.find_elements(By.XPATH, "//div[@class='_2a-B5 VfC3c' and contains(text(), 'Contacts on WhatsApp')]")
    not_in_contact = driver.find_elements(By.XPATH, "//div[@class='_2a-B5 VfC3c' and contains(text(), 'Not in your contacts')]")
    not_in_whatsapp = driver.find_elements(By.XPATH, "//span[@class='_11JPr' and contains(text(), 'No results found')]")
    time.sleep(0.7)
    if in_contact:
        button = driver.find_element(By.XPATH, "//div[@tabindex='-1' and @role='button']")
        button.click()
    elif not_in_contact:
        button = driver.find_element(By.XPATH, "//div[@class='g0rxnol2 g0rxnol2 thghmljt p357zi0d rjo8vgbg ggj6brxn f8m0rgwh gfz4du6o ag5g9lrv bs7a17vp ov67bkzj']/div[@class='_2a-B5 VfC3c' and contains(text(), 'Not in your contacts')]/following-sibling::div[@tabindex='0' and @role='button']")
        button.click()
    elif not_in_whatsapp:
        cancel_button = driver.find_element(By.XPATH, "//span[contains(@data-icon, 'x-alt')]")
        cancel_button.click()
        logging.warning(f"{num} is not on WhatsApp")
        return

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@contenteditable='true' and @data-tab='1']")))

    # Replace placeholders in the message with the corresponding values
    personalized_message = original_message.replace('{{query_name}}', name).replace('{{query_city}}', city_name)

    # Using pyperclip to avoid issues with send_keys and formatting
    pyperclip.copy(personalized_message)
    input_box = driver.find_element(By.XPATH, "//div[@contenteditable='true' and @data-tab='1']")
    input_box.click()
    input_box.send_keys(Keys.CONTROL + 'v')  # Paste the message from clipboard
    input_box.send_keys(Keys.SHIFT, Keys.ENTER)  # Press shift+enter to avoid immediate send
    
    # Attach file if provided
    if attached_document:
        attach_button = driver.find_element(By.XPATH, "//div[@title='Attach']")
        attach_button.click()
        time.sleep(0.5)

        # Determine file type based on extension
        if attached_document.endswith(('.png', '.jpg', '.jpeg', '.gif')):
            media_type_button = driver.find_element(By.XPATH, "//input[@accept='image/*,video/mp4,video/3gpp,video/quicktime']")
        elif attached_document.endswith(('.mp4', '.avi', '.mov')):
            media_type_button = driver.find_element(By.XPATH, "//input[@accept='video/mp4,video/3gpp,video/quicktime']")
        else:
            raise ValueError("Unsupported file type for attachment.")

        media_type_button.send_keys(attached_document)
        time.sleep(2)

    # Send message
    send_button = driver.find_element(By.XPATH, "//span[@data-icon='send']")
    send_button.click()
    time.sleep(random.randint(3, 5))  # Add random delay between messages to mimic human behavior

def delete_message():
    numbers_entry.delete("1.0", "end")
    message_entry.delete("1.0", "end")

def delete_attachment():
    global attached_document
    attached_document = ''
    attachment_label.config(text="No attachment")
    delete_attachment.config(state=tk.DISABLED)

# Creating the UI
window = tk.Tk()
window.title("WhatsApp Bulk Messenger")
window.geometry("500x550")

# Frame for labels and entry boxes
frame = tk.Frame(window)
frame.pack(pady=10)

# Numbers input
numbers_label = tk.Label(frame, text="Phone Numbers:")
numbers_label.grid(row=0, column=0, sticky=tk.W)
numbers_entry = tk.Text(frame, height=5, width=50)
numbers_entry.grid(row=1, column=0, columnspan=2, padx=10, pady=10)

# Upload Excel button
upload_button = tk.Button(frame, text="Upload Excel", command=upload_excel)
upload_button.grid(row=2, column=0, sticky=tk.W, padx=10)

# Clear button
delete_button = tk.Button(frame, text="Clear", state=tk.DISABLED, command=delete_message)
delete_button.grid(row=2, column=1, sticky=tk.E, padx=10)

# Message input
message_label = tk.Label(frame, text="Message:")
message_label.grid(row=3, column=0, sticky=tk.W)
message_entry = tk.Text(frame, height=10, width=50)
message_entry.grid(row=4, column=0, columnspan=2, padx=10, pady=10)

# Insert query placeholders
name_button = tk.Button(frame, text="Insert Name", state=tk.DISABLED, command=insert_query_name)
name_button.grid(row=5, column=0, padx=10, sticky=tk.W)
city_button = tk.Button(frame, text="Insert City", state=tk.DISABLED, command=insert_query_city)
city_button.grid(row=5, column=1, padx=10, sticky=tk.E)

# Attach file button
attach_button = tk.Button(frame, text="Attach File", command=attach_file)
attach_button.grid(row=6, column=0, padx=10, pady=5, sticky=tk.W)
attachment_label = tk.Label(frame, text="No attachment")
attachment_label.grid(row=6, column=1, sticky=tk.E)
delete_attachment = tk.Button(frame, text="Remove Attachment", state=tk.DISABLED, command=delete_attachment)
delete_attachment.grid(row=6, column=2, padx=10, pady=5, sticky=tk.E)

# Progress bar
progress_var = tk.DoubleVar()
progress_bar = ttk.Progressbar(window, variable=progress_var, maximum=100)
progress_bar.pack(pady=10)

# Sent messages count
sent_label = tk.Label(window, text="Messages sent: 0/0")
sent_label.pack()

# Send button
send_button = tk.Button(window, text="Send Messages", command=send_messages_async)
send_button.pack(pady=20)

# Start the application
window.mainloop()

